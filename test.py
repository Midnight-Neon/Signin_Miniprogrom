from openpyxl import load_workbook, Workbook
import copy

wb = load_workbook("导入.xlsx")
ws = wb.worksheets[0]
rows = ws.max_row
columns = ws.max_column
print(rows, columns)
if columns < 7:
    print("Error!")
    exit(-1)
info = {}
infos = []
cata = {}
for row in ws.iter_rows(min_row=2):
    infox = copy.deepcopy(info)
    infox['ID'] = str(row[0].value)
    infox['_id'] = str(row[0].value)

    infox['PassWd'] = '123456'
    infox['name'] = row[1].value
    infox['department'] = row[2].value
    infox['grades'] = row[3].value
    infox['major'] = row[4].value
    infox['group'] = row[5].value
    if infox['department'] in cata:
        if infox['major'] in cata.get(infox['department'], {}):
            if infox['group'] not in cata.get(infox['department'], {}).get(infox['major'], []):
                cata[infox['department']][infox['major']].append(infox['group'])
        else:
            cata[infox['department']][infox['major']] = []
            cata[infox['department']][infox['major']].append(infox['group'])
    else:
        cata[infox['department']] = {}
        cata[infox['department']][infox['major']] = []
        cata[infox['department']][infox['major']].append(infox['group'])
    if row[6].value is None:
        infox['role'] = 0
    else:
        infox['role'] = 1

    infos.append(infox)
print(infos)
print(cata)
